import json
from datetime import date, datetime, timedelta
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Sum, Count
from django.utils import timezone
from django.template.loader import render_to_string
from obat.models import Barang, StokBarang, DaftarHarga
from .models import Pembelian, LaporanHarian
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


# ─── PEMBELIAN ────────────────────────────────────────────
@csrf_exempt
def pembelian_list(request):
    if request.method == 'GET':
        data = list(Pembelian.objects.values(
            'id', 'barang_id', 'nama_barang', 'jumlah_pembelian',
            'harga', 'total_pembayaran', 'tgl_beli'
        ).order_by('-tgl_beli'))
        for d in data:
            d['harga'] = float(d['harga'])
            d['total_pembayaran'] = float(d['total_pembayaran'])
            d['tgl_beli'] = str(d['tgl_beli'])
        return JsonResponse({'status': 'ok', 'data': data})

    if request.method == 'POST':
        body = json.loads(request.body)
        id_barang = body['id_barang']

        try:
            barang = Barang.objects.get(pk=id_barang)
            stok = StokBarang.objects.get(barang=barang)
            harga_obj = DaftarHarga.objects.get(barang=barang)
        except (Barang.DoesNotExist, StokBarang.DoesNotExist, DaftarHarga.DoesNotExist):
            return JsonResponse({'status': 'error', 'message': 'Barang tidak ditemukan'}, status=404)

        jumlah = int(body['jumlah_pembelian'])
        if stok.jumlah_stok < jumlah:
            return JsonResponse({'status': 'error', 'message': f'Stok tidak cukup. Stok tersedia: {stok.jumlah_stok}'}, status=400)

        harga = float(harga_obj.harga_satuan)
        total = harga * jumlah

        p = Pembelian.objects.create(
            barang=barang,
            nama_barang=barang.nama_barang,
            jumlah_pembelian=jumlah,
            harga=harga,
            total_pembayaran=total,
        )

        # Kurangi stok otomatis
        stok.jumlah_stok -= jumlah
        stok.save()
        barang.jumlah_barang = stok.jumlah_stok
        barang.save()

        return JsonResponse({
            'status': 'ok',
            'message': 'Pembelian berhasil',
            'data': {
                'id': p.id,
                'nama_barang': p.nama_barang,
                'jumlah': jumlah,
                'harga_satuan': harga,
                'total': total
            }
        }, status=201)


@csrf_exempt
def pembelian_detail(request, pk):
    try:
        p = Pembelian.objects.get(pk=pk)
    except Pembelian.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Data tidak ditemukan'}, status=404)

    if request.method == 'GET':
        return JsonResponse({'status': 'ok', 'data': {
            'id': p.id, 'id_barang': p.barang_id,
            'nama_barang': p.nama_barang, 'jumlah_pembelian': p.jumlah_pembelian,
            'harga': float(p.harga), 'total_pembayaran': float(p.total_pembayaran),
            'tgl_beli': str(p.tgl_beli)
        }})

    if request.method == 'DELETE':
        p.delete()
        return JsonResponse({'status': 'ok', 'message': 'Data pembelian dihapus'})


# ─── SINKRONISASI LAPORAN HARIAN ─────────────────────────
@csrf_exempt
def sinkronisasi_laporan(request):
    if request.method == 'POST':
        body = json.loads(request.body) if request.body else {}
        tgl = body.get('tgl_laporan', str(date.today()))

        tgl_date = datetime.strptime(tgl, '%Y-%m-%d').date()
        tgl_start = timezone.make_aware(datetime.combine(tgl_date, datetime.min.time()))
        tgl_end = timezone.make_aware(datetime.combine(tgl_date + timedelta(days=1), datetime.min.time()))

        total_transaksi = Pembelian.objects.filter(tgl_beli__gte=tgl_start, tgl_beli__lt=tgl_end).count()
        total_pendapatan = Pembelian.objects.filter(tgl_beli__gte=tgl_start, tgl_beli__lt=tgl_end).aggregate(
            total=Sum('total_pembayaran')
        )['total'] or 0

        laporan, created = LaporanHarian.objects.update_or_create(
            tgl_laporan=tgl_date,
            defaults={
                'total_transaksi': total_transaksi,
                'total_pendapatan': total_pendapatan,
            }
        )

        return JsonResponse({
            'status': 'ok',
            'message': f'Laporan {"dibuat" if created else "diperbarui"} — {total_transaksi} transaksi, Rp {float(total_pendapatan):,.0f}',
            'data': {
                'tgl_laporan': str(laporan.tgl_laporan),
                'total_transaksi': laporan.total_transaksi,
                'total_pendapatan': float(laporan.total_pendapatan),
            }
        })

    if request.method == 'GET':
        data = list(LaporanHarian.objects.values(
            'id', 'tgl_laporan', 'total_transaksi', 'total_pendapatan', 'generated_at'
        ).order_by('-tgl_laporan'))
        for d in data:
            d['tgl_laporan'] = str(d['tgl_laporan'])
            d['total_pendapatan'] = float(d['total_pendapatan'])
            d['generated_at'] = str(d['generated_at'])
        return JsonResponse({'status': 'ok', 'data': data})


# ─── EXPORT LAPORAN KE EXCEL ──────────────────────────────
@csrf_exempt
def export_laporan_excel(request):
    laporan_list = LaporanHarian.objects.all().order_by('-tgl_laporan')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Laporan Harian'

    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='2563A8', end_color='2563A8', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )

    headers = ['No', 'Tanggal', 'Total Transaksi', 'Total Pendapatan (Rp)', 'Terakhir Diperbarui']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    data_font = Font(name='Calibri', size=11)
    data_align = Alignment(horizontal='center', vertical='center')
    rupiah_align = Alignment(horizontal='right', vertical='center')

    for i, l in enumerate(laporan_list, 1):
        row = i + 1
        ws.cell(row=row, column=1, value=i).font = data_font
        ws.cell(row=row, column=1).alignment = data_align
        ws.cell(row=row, column=1).border = thin_border

        ws.cell(row=row, column=2, value=str(l.tgl_laporan)).font = data_font
        ws.cell(row=row, column=2).alignment = data_align
        ws.cell(row=row, column=2).border = thin_border

        ws.cell(row=row, column=3, value=l.total_transaksi).font = data_font
        ws.cell(row=row, column=3).alignment = data_align
        ws.cell(row=row, column=3).border = thin_border

        ws.cell(row=row, column=4, value=float(l.total_pendapatan)).font = data_font
        ws.cell(row=row, column=4).alignment = rupiah_align
        ws.cell(row=row, column=4).number_format = '#,##0'
        ws.cell(row=row, column=4).border = thin_border

        ws.cell(row=row, column=5, value=l.generated_at.strftime('%Y-%m-%d %H:%M')).font = data_font
        ws.cell(row=row, column=5).alignment = data_align
        ws.cell(row=row, column=5).border = thin_border

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 24
    ws.column_dimensions['E'].width = 22

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="laporan-harian-{date.today()}.xlsx"'
    wb.save(response)
    return response


# ─── EXPORT LAPORAN KE PDF ────────────────────────────────
@csrf_exempt
def export_laporan_pdf(request):
    laporan_list = LaporanHarian.objects.all().order_by('-tgl_laporan')
    total_pendapatan_all = sum(float(l.total_pendapatan) for l in laporan_list)
    total_transaksi_all = sum(l.total_transaksi for l in laporan_list)

    html = render_to_string('laporan_pdf.html', {
        'laporan_list': laporan_list,
        'total_pendapatan_all': total_pendapatan_all,
        'total_transaksi_all': total_transaksi_all,
        'today': datetime.now(),
    })
    return HttpResponse(html)
